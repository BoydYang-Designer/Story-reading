import os
import openpyxl
from openpyxl.styles import Alignment
from tkinter import Tk, messagebox, Toplevel, Label, Button, Radiobutton, StringVar, Entry, Listbox, Scrollbar, SINGLE, Frame
from tkinter.filedialog import askopenfilename, askopenfilenames, askdirectory
import re
import difflib  # 新增：用來顯示內容差異


def extract_category_and_title_from_filename(filename):
    """
    從檔名提取分類和標題
    """
    name_without_ext = os.path.splitext(filename)[0]
    title = name_without_ext.strip()
    
    if " - " in name_without_ext:
        category = name_without_ext.split(" - ")[0].strip()
    elif " " in name_without_ext:
        category = name_without_ext.split(" ")[0].strip()
    elif "_" in name_without_ext:
        category = name_without_ext.split("_")[0].strip()
    elif "-" in name_without_ext:
        category = name_without_ext.split("-")[0].strip()
    else:
        category = name_without_ext.strip()
    
    return category, title


def clean_timestamp_content(content: str) -> str:
    """
    專門針對 Timestamp 字幕檔移除時間戳記
    範例: [00:00:00.000 --> 00:00:00.960] Well, usually... → Well, usually...
    只保留真正文字內容，並去除空白行
    """
    lines = content.splitlines()
    cleaned_lines = []
    
    # 匹配常見字幕格式（含空格）
    timestamp_pattern = re.compile(r'^\s*\[\d{2}:\d{2}:\d{2}\.\d{3}\s*-->\s*\d{2}:\d{2}:\d{2}\.\d{3}\]\s*')
    
    for line in lines:
        # 移除時間戳記
        clean_line = timestamp_pattern.sub('', line).strip()
        if clean_line:  # 只保留有內容的行
            cleaned_lines.append(clean_line)
    
    return '\n'.join(cleaned_lines)


def get_existing_categories(sheet):
    """
    從 Excel 中獲取所有已存在的大類（A 欄）
    """
    categories = set()
    for row in range(2, sheet.max_row + 1):
        category = sheet.cell(row=row, column=1).value
        if category and isinstance(category, str) and category.strip():
            categories.add(category.strip())
    return sorted(list(categories))


class CategoryDialog:
    # （完全不變，維持原功能）
    def __init__(self, parent, existing_categories):
        self.result = None
        self.mode = None
        self.dialog = Toplevel(parent)
        self.dialog.title("大類設定")
        self.dialog.resizable(False, False)
        
        self.dialog.update_idletasks()
        screen_width = self.dialog.winfo_screenwidth()
        screen_height = self.dialog.winfo_screenheight()
        x = (screen_width - 500) // 2
        y = (screen_height - 450) // 2
        self.dialog.geometry(f"500x450+{x}+{y}")
        
        self.dialog.lift()
        self.dialog.focus_force()
        self.dialog.grab_set()
        self.dialog.attributes('-topmost', True)
        self.dialog.update()
        self.dialog.attributes('-topmost', False)
        
        Label(self.dialog, text="設定大類（A 欄）", 
              font=('微軟正黑體', 12, 'bold')).pack(pady=10)
        Label(self.dialog, text="請選擇如何處理新增資料的大類：", 
              font=('微軟正黑體', 10)).pack(pady=5)
        
        self.choice = StringVar(value="blank")
        
        Radiobutton(self.dialog, text="保持空白", variable=self.choice, 
                   value="blank", font=('微軟正黑體', 10)).pack(anchor='w', padx=30, pady=5)
        Radiobutton(self.dialog, text="自訂大類（手動輸入）", variable=self.choice, 
                   value="custom", font=('微軟正黑體', 10)).pack(anchor='w', padx=30, pady=5)
        
        self.custom_entry = Entry(self.dialog, font=('微軟正黑體', 10), width=30)
        self.custom_entry.pack(padx=50, pady=5)
        self.custom_entry.config(state='disabled')
        
        if existing_categories:
            Radiobutton(self.dialog, text="從現有大類中選擇", variable=self.choice, 
                       value="existing", font=('微軟正黑體', 10)).pack(anchor='w', padx=30, pady=5)
            
            list_frame = Frame(self.dialog)
            list_frame.pack(padx=50, pady=5)
            Label(list_frame, text="現有大類：", font=('微軟正黑體', 9)).pack(anchor='w')
            
            scrollbar = Scrollbar(list_frame)
            scrollbar.pack(side='right', fill='y')
            
            self.category_listbox = Listbox(list_frame, height=6, width=35, 
                                           font=('微軟正黑體', 10), 
                                           yscrollcommand=scrollbar.set,
                                           selectmode=SINGLE)
            self.category_listbox.pack(side='left')
            scrollbar.config(command=self.category_listbox.yview)
            
            for cat in existing_categories:
                self.category_listbox.insert('end', cat)
            if existing_categories:
                self.category_listbox.select_set(0)
            self.category_listbox.config(state='disabled')
        else:
            self.category_listbox = None
        
        self.choice.trace('w', self.on_choice_change)
        
        button_frame = Frame(self.dialog)
        button_frame.pack(pady=20)
        Button(button_frame, text="確定", command=self.on_ok, 
               width=10, font=('微軟正黑體', 10)).pack(side='left', padx=10)
        Button(button_frame, text="取消", command=self.on_cancel, 
               width=10, font=('微軟正黑體', 10)).pack(side='left', padx=10)
        
        self.dialog.wait_window()
    
    def on_choice_change(self, *args):
        choice = self.choice.get()
        if choice == "custom":
            self.custom_entry.config(state='normal')
        else:
            self.custom_entry.config(state='disabled')
        
        if self.category_listbox:
            if choice == "existing":
                self.category_listbox.config(state='normal')
            else:
                self.category_listbox.config(state='disabled')
    
    def on_ok(self):
        choice = self.choice.get()
        if choice == "blank":
            self.result = None
            self.mode = "blank"
        elif choice == "custom":
            custom_text = self.custom_entry.get().strip()
            if not custom_text:
                messagebox.showwarning("警告", "請輸入自訂大類，或選擇其他選項")
                return
            self.result = custom_text
            self.mode = "custom"
        elif choice == "existing":
            if self.category_listbox:
                selection = self.category_listbox.curselection()
                if not selection:
                    messagebox.showwarning("警告", "請選擇一個現有大類，或選擇其他選項")
                    return
                self.result = self.category_listbox.get(selection[0])
                self.mode = "existing"
        
        self.dialog.destroy()
    
    def on_cancel(self):
        self.result = None
        self.mode = None
        self.dialog.destroy()


def scan_folder_for_txt(folder_path):
    """
    掃描資料夾中的所有 txt 檔案（包含 Timestamp 檔案）
    """
    txt_files = []
    for file in os.listdir(folder_path):
        file_path = os.path.join(folder_path, file)
        if os.path.isfile(file_path):
            _, ext = os.path.splitext(file)
            if ext.lower() == '.txt':
                txt_files.append(file_path)
    return sorted(txt_files)


def find_next_empty_row(sheet, start_row=2):
    """
    找到下一個可以填入的空白行（B 欄和 C 欄都為空）
    """
    current_row = start_row
    max_row = sheet.max_row
    while current_row <= max_row + 1:
        cell_b = sheet.cell(row=current_row, column=2).value
        cell_c = sheet.cell(row=current_row, column=3).value
        if cell_b is None and cell_c is None:
            return current_row
        current_row += 1
    return current_row


def find_title_row(sheet, title):
    """
    檢查標題是否已經存在於 Excel 中（C 欄），回傳行號或 None
    """
    for row in range(2, sheet.max_row + 1):
        cell_value = sheet.cell(row=row, column=3).value
        if cell_value and cell_value == title:
            return row
    return None


def fill_excel_with_txt(excel_path, txt_files, major_category=None):
    """
    將 txt 檔案內容填入 Excel
    新增功能：
    1. Timestamp 檔案自動移除時間戳記
    2. 若標題已存在且 D 欄內容不同 → 顯示差異並詢問是否覆蓋
    """
    try:
        print(f"正在開啟 Excel 檔案: {excel_path}")
        wb = openpyxl.load_workbook(excel_path)
        sheet = wb.active
        
        success_count = 0      # 新增 + 成功覆蓋
        skip_same_count = 0    # 內容完全相同，跳過
        
        for txt_file in txt_files:
            try:
                filename = os.path.basename(txt_file)
                print(f"\n處理檔案: {filename}")
                
                # 讀取原始內容
                with open(txt_file, 'r', encoding='utf-8') as f:
                    content_raw = f.read().strip()
                
                # === 新增：Timestamp 檔案自動清理時間戳記 ===
                is_timestamp_file = 'timestamp' in filename.lower()
                if is_timestamp_file:
                    content = clean_timestamp_content(content_raw)
                    print(f"  → Timestamp 檔案，已自動移除時間戳記")
                else:
                    content = content_raw
                
                # 提取分類與標題
                category, title = extract_category_and_title_from_filename(filename)
                print(f"  提取的分類: {category}")
                print(f"  提取的標題: {title}")
                
                # === 新增：檢查標題是否已存在 + 內容比對 ===
                existing_row = find_title_row(sheet, title)
                
                if existing_row:
                    existing_content = sheet.cell(row=existing_row, column=4).value or ""
                    
                    # 忽略純空白差異
                    if existing_content.strip() == content.strip():
                        print(f"  ✓ 標題 '{title}' 已存在且內容完全相同，跳過")
                        skip_same_count += 1
                        continue
                    
                    # 內容不同 → 顯示差異並詢問
                    diff_lines = list(difflib.unified_diff(
                        existing_content.splitlines(keepends=True),
                        content.splitlines(keepends=True),
                        fromfile='現有 Excel 內容',
                        tofile='新 TXT 內容 (已清理)',
                        lineterm=''
                    ))
                    diff_text = ''.join(diff_lines) if diff_lines else "無明顯差異（僅空白或換行）"
                    
                    # 防止訊息框過長
                    if len(diff_text) > 2500:
                        diff_text = diff_text[:2500] + "\n\n...（差異過長，已截斷）"
                    
                    confirm_overwrite = messagebox.askyesno(
                        "內容差異偵測 - 是否覆蓋？",
                        f"標題「{title}」已存在於 Excel 第 {existing_row} 行。\n\n"
                        f"D 欄內容與新 TXT 檔案不相同！\n\n"
                        f"差異預覽（unified diff）：\n\n{diff_text}\n\n"
                        f"是否要覆蓋 Excel 中的 D 欄內容？\n"
                        f"（A、B、C 欄不會改變）"
                    )
                    
                    if not confirm_overwrite:
                        print(f"  ✗ 使用者選擇不覆蓋，跳過")
                        skip_same_count += 1
                        continue
                    
                    target_row = existing_row
                    print(f"  → 使用者同意覆蓋第 {target_row} 行的 D 欄內容")
                
                else:
                    # 新標題 → 找下一行
                    target_row = find_next_empty_row(sheet)
                    print(f"  填入新位置: 第 {target_row} 行")
                
                # === 填入資料（新舊都走同一段）===
                if major_category:
                    sheet.cell(row=target_row, column=1).value = major_category
                else:
                    sheet.cell(row=target_row, column=1).value = None
                
                sheet.cell(row=target_row, column=2).value = category
                
                cell_c = sheet.cell(row=target_row, column=3)
                cell_c.value = title
                
                cell_d = sheet.cell(row=target_row, column=4)
                cell_d.value = content
                cell_d.alignment = Alignment(wrap_text=True, vertical='top')
                
                # E 欄公式自動調整（與原版相同）
                formula_pattern = None
                for check_row in range(2, min(10, sheet.max_row + 1)):
                    check_cell = sheet.cell(row=check_row, column=5)
                    if check_cell.value and isinstance(check_cell.value, str) and check_cell.value.startswith('=HYPERLINK'):
                        formula_pattern = check_cell.value
                        break
                if formula_pattern:
                    new_formula = re.sub(r'C\d+', f'C{target_row}', formula_pattern)
                    sheet.cell(row=target_row, column=5).value = new_formula
                
                print(f"  ✓ 成功填入/更新")
                success_count += 1
                
            except Exception as e:
                print(f"  ✗ 處理檔案時發生錯誤: {e}")
                continue
        
        # 儲存
        print(f"\n正在儲存 Excel 檔案...")
        wb.save(excel_path)
        print(f"✓ Excel 檔案已更新")
        
        return success_count, skip_same_count
        
    except Exception as e:
        print(f"發生錯誤: {e}")
        raise


def main():
    """
    主程式 - v3.2（支援 Timestamp 檔案匯入 + 自動移除時間戳記 + 內容差異覆蓋確認）
    """
    print("=" * 60)
    print("Excel 自動填入工具 v3.2（Timestamp 增強版）")
    print("=" * 60)
    
    root = Tk()
    root.withdraw()
    
    # 步驟 1: 選擇 Excel
    print("\n請選擇要更新的 Excel 檔案...")
    excel_path = askopenfilename(
        title="選擇要更新的 Excel 檔案",
        filetypes=[("Excel 檔案", "*.xlsx"), ("所有檔案", "*.*")]
    )
    
    if not excel_path:
        print("未選擇 Excel 檔案，程式結束。")
        messagebox.showinfo("提示", "未選擇 Excel 檔案")
        return
    
    print(f"已選擇 Excel: {os.path.basename(excel_path)}")
    
    # 步驟 2: 選擇處理模式
    mode_choice = messagebox.askquestion(
        "選擇處理模式",
        "請選擇 txt 檔案來源:\n\n"
        "● Yes: 手動選擇多個 txt 檔案\n"
        "● No: 選擇資料夾（自動掃描所有 txt，包含 Timestamp）"
    )
    
    if mode_choice == "yes":
        print("\n請選擇要匯入的 txt 檔案（可複選）...")
        txt_files = list(askopenfilenames(
            title="選擇 txt 檔案（可複選）",
            filetypes=[("文字檔案", "*.txt"), ("所有檔案", "*.*")]
        ))
        if not txt_files:
            print("未選擇 txt 檔案，程式結束。")
            messagebox.showinfo("提示", "未選擇 txt 檔案")
            return
        print(f"已選擇 {len(txt_files)} 個 txt 檔案")
    else:
        print("\n請選擇包含 txt 檔案的資料夾...")
        folder_path = askdirectory(title="選擇包含 txt 檔案的資料夾")
        if not folder_path:
            print("未選擇資料夾，程式結束。")
            messagebox.showinfo("提示", "未選擇資料夾")
            return
        txt_files = scan_folder_for_txt(folder_path)
        if not txt_files:
            print("資料夾中沒有找到 txt 檔案。")
            messagebox.showinfo("提示", "資料夾中沒有找到 txt 檔案")
            return
        print(f"在資料夾中找到 {len(txt_files)} 個 txt 檔案（含 Timestamp）")
    
    # 步驟 3: 大類設定（完全不變）
    print("\n正在讀取 Excel 中的現有大類...")
    try:
        wb = openpyxl.load_workbook(excel_path)
        sheet = wb.active
        existing_categories = get_existing_categories(sheet)
        wb.close()
        if existing_categories:
            print(f"找到 {len(existing_categories)} 個現有大類")
        else:
            print("Excel 中尚無現有大類")
    except Exception as e:
        print(f"讀取 Excel 失敗: {e}")
        messagebox.showerror("錯誤", f"無法讀取 Excel 檔案：\n{e}")
        return
    
    category_dialog = CategoryDialog(root, existing_categories)
    if category_dialog.mode is None:
        print("使用者取消大類設定")
        return
    
    major_category = category_dialog.result
    print(f"大類設定: {major_category if major_category else '(保持空白)'}")
    
    # 確認操作（已更新說明）
    confirm_message = (
        f"即將處理：\n"
        f"• Excel 檔案: {os.path.basename(excel_path)}\n"
        f"• txt 檔案數量: {len(txt_files)}\n\n"
        f"填入規則：\n"
        f"• 大類（A欄）：{major_category if major_category else '(保持空白)'}\n"
        f"• 分類（B欄）：檔名的第一個單字\n"
        f"• 標題（C欄）：完整檔名\n"
        f"• 內文（D欄）：txt 內容（自動換行）\n"
        f"• Timestamp 檔案：自動移除時間戳記後匯入\n"
        f"• 若標題已存在且 D 欄內容不同：顯示差異並詢問是否覆蓋\n\n"
        f"Excel 檔案將被直接覆蓋更新。\n"
        f"是否繼續？"
    )
    
    confirm = messagebox.askyesno("確認操作", confirm_message)
    if not confirm:
        print("使用者取消操作")
        return
    
    # 步驟 4: 處理檔案
    try:
        print("\n" + "=" * 60)
        print("開始處理...")
        print("=" * 60)
        
        success_count, skip_same_count = fill_excel_with_txt(
            excel_path, txt_files, major_category
        )
        
        print("\n" + "=" * 60)
        print("處理完成！")
        print(f"成功填入/更新: {success_count} 個")
        print(f"跳過（內容相同）: {skip_same_count} 個")
        print("=" * 60)
        
        result_message = (
            f"處理完成！\n\n"
            f"成功填入/更新: {success_count} 個\n"
            f"跳過（內容相同）: {skip_same_count} 個\n\n"
            f"Excel 檔案已更新：\n{excel_path}"
        )
        messagebox.showinfo("完成", result_message)
        
    except Exception as e:
        error_msg = f"處理過程發生錯誤：\n{str(e)}"
        print(f"\n錯誤: {error_msg}")
        messagebox.showerror("錯誤", error_msg)


if __name__ == "__main__":
    main()