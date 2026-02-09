import os
import openpyxl
from openpyxl.styles import Alignment
from tkinter import Tk, messagebox
from tkinter.filedialog import askopenfilename, askopenfilenames, askdirectory
import re

def extract_category_and_title_from_filename(filename):
    """
    從檔名提取分類和標題
    例如: "Elicitation - How to Get People to Talk Without Them Realizing.txt"
    分類: "Elicitation"
    標題: "Elicitation - How to Get People to Talk Without Them Realizing"
    """
    # 移除 .txt 副檔名
    name_without_ext = os.path.splitext(filename)[0]
    
    # 標題就是完整的檔名（不含副檔名）
    title = name_without_ext.strip()
    
    # 提取第一個單字作為分類
    # 如果有 " - " 分隔符，取第一部分
    if " - " in name_without_ext:
        category = name_without_ext.split(" - ")[0].strip()
    # 如果有空格，取第一個單字
    elif " " in name_without_ext:
        category = name_without_ext.split(" ")[0].strip()
    # 如果有底線，取第一部分
    elif "_" in name_without_ext:
        category = name_without_ext.split("_")[0].strip()
    # 如果有減號（但沒有空格），取第一部分
    elif "-" in name_without_ext:
        category = name_without_ext.split("-")[0].strip()
    else:
        # 如果沒有任何分隔符，使用整個檔名作為分類
        category = name_without_ext.strip()
    
    return category, title

def scan_folder_for_txt(folder_path):
    """
    掃描資料夾中的所有 txt 檔案
    排除檔名中包含 "Timestamp" 的檔案
    
    參數:
    - folder_path: 資料夾路徑
    
    返回:
    - txt 檔案路徑列表
    - 被跳過的 Timestamp 檔案列表
    """
    txt_files = []
    skipped_timestamp_files = []
    
    for file in os.listdir(folder_path):
        file_path = os.path.join(folder_path, file)
        
        # 只處理檔案（非資料夾）
        if os.path.isfile(file_path):
            # 檢查副檔名是否為 .txt
            _, ext = os.path.splitext(file)
            if ext.lower() == '.txt':
                # 檢查檔名是否包含 "Timestamp"
                if 'Timestamp' in file or 'timestamp' in file:
                    skipped_timestamp_files.append(file)
                else:
                    txt_files.append(file_path)
    
    return sorted(txt_files), sorted(skipped_timestamp_files)

def find_next_empty_row(sheet, start_row=2):
    """
    找到下一個可以填入的空白行
    從 start_row 開始尋找，當 B 欄和 C 欄都為空時返回該行號
    """
    current_row = start_row
    max_row = sheet.max_row
    
    # 從現有資料後面開始
    while current_row <= max_row + 1:
        cell_b = sheet.cell(row=current_row, column=2).value  # B 欄（分類）
        cell_c = sheet.cell(row=current_row, column=3).value  # C 欄（標題）
        
        # 如果 B 欄和 C 欄都是空的，這就是我們要的空白行
        if cell_b is None and cell_c is None:
            return current_row
        
        current_row += 1
    
    return current_row

def check_title_exists(sheet, title):
    """
    檢查標題是否已經存在於 Excel 中（C 欄）
    """
    for row in range(2, sheet.max_row + 1):
        cell_value = sheet.cell(row=row, column=3).value
        if cell_value and cell_value == title:
            return True
    return False

def fill_excel_with_txt(excel_path, txt_files):
    """
    將 txt 檔案內容填入 Excel
    """
    try:
        # 載入 Excel 檔案
        print(f"正在開啟 Excel 檔案: {excel_path}")
        wb = openpyxl.load_workbook(excel_path)
        sheet = wb.active
        
        success_count = 0
        skip_exist_count = 0
        skip_timestamp_count = 0
        
        for txt_file in txt_files:
            try:
                # 讀取 txt 檔案內容
                filename = os.path.basename(txt_file)
                
                # 檢查檔名是否包含 "Timestamp"，如果是則跳過
                if 'Timestamp' in filename or 'timestamp' in filename:
                    print(f"\n✗ 跳過 Timestamp 檔案: {filename}")
                    skip_timestamp_count += 1
                    continue
                
                print(f"\n處理檔案: {filename}")
                
                with open(txt_file, 'r', encoding='utf-8') as f:
                    content = f.read().strip()
                
                # 提取分類和標題
                category, title = extract_category_and_title_from_filename(filename)
                print(f"  提取的分類: {category}")
                print(f"  提取的標題: {title}")
                
                # 檢查標題是否已存在
                if check_title_exists(sheet, title):
                    print(f"  ✗ 跳過：標題 '{title}' 已存在於 Excel 中")
                    skip_exist_count += 1
                    continue
                
                # 找到下一個空白行
                target_row = find_next_empty_row(sheet)
                print(f"  填入位置: 第 {target_row} 行")
                
                # 填入資料
                # A 欄（大類）保持空白
                sheet.cell(row=target_row, column=1).value = None
                
                # B 欄（分類）填入提取的分類
                sheet.cell(row=target_row, column=2).value = category
                
                # C 欄（標題）填入完整檔名（不含副檔名）
                cell_c = sheet.cell(row=target_row, column=3)
                cell_c.value = title
                
                # D 欄（內文）填入 txt 檔案內容，並設定自動換行
                cell_d = sheet.cell(row=target_row, column=4)
                cell_d.value = content
                cell_d.alignment = Alignment(wrap_text=True, vertical='top')
                
                # E 欄（如果有公式，保持原有格式）
                # 檢查其他行的 E 欄是否有公式模式
                formula_pattern = None
                for check_row in range(2, min(10, sheet.max_row + 1)):
                    check_cell = sheet.cell(row=check_row, column=5)
                    if check_cell.value and isinstance(check_cell.value, str) and check_cell.value.startswith('=HYPERLINK'):
                        # 找到公式模式，替換行號
                        formula_pattern = check_cell.value
                        break
                
                if formula_pattern:
                    # 將公式中的行號替換為當前行
                    # 假設公式格式為: =HYPERLINK("audio\" & C2 & ".mp3", C2)
                    new_formula = re.sub(r'C\d+', f'C{target_row}', formula_pattern)
                    sheet.cell(row=target_row, column=5).value = new_formula
                
                print(f"  ✓ 成功填入")
                success_count += 1
                
            except Exception as e:
                print(f"  ✗ 處理檔案時發生錯誤: {e}")
                continue
        
        # 儲存 Excel 檔案（覆蓋原檔）
        print(f"\n正在儲存 Excel 檔案...")
        wb.save(excel_path)
        print(f"✓ Excel 檔案已更新")
        
        return success_count, skip_exist_count, skip_timestamp_count
        
    except Exception as e:
        print(f"發生錯誤: {e}")
        raise

def main():
    """
    主程式
    """
    print("=" * 60)
    print("Excel 自動填入工具 v3.0")
    print("=" * 60)
    
    # 初始化 Tkinter
    root = Tk()
    root.withdraw()
    
    # 步驟 1: 選擇要更新的 Excel 檔案
    print("\n請選擇要更新的 Excel 檔案...")
    excel_path = askopenfilename(
        title="選擇要更新的 Excel 檔案",
        filetypes=[("Excel 檔案", "*.xlsx"), ("所有檔案", "*.*")]
    )
    
    if not excel_path:
        print("未選擇 Excel 檔案，程式結束。")
        messagebox.showinfo("提示", "未選擇 Excel 檔案")
        return
    
    print(f"已選擇 Excel: {os.path.basename(excel_path)}")
    
    # 步驟 2: 選擇處理模式（選擇檔案 or 選擇資料夾）
    mode_choice = messagebox.askquestion(
        "選擇處理模式",
        "請選擇 txt 檔案來源:\n\n"
        "● Yes: 手動選擇多個 txt 檔案\n"
        "● No: 選擇資料夾（自動掃描所有 txt）"
    )
    
    txt_files = []
    skipped_timestamp_files = []
    
    if mode_choice == "yes":
        # 手動選擇多個 txt 檔案
        print("\n請選擇要匯入的 txt 檔案（可複選）...")
        txt_files = list(askopenfilenames(
            title="選擇 txt 檔案（可複選）",
            filetypes=[("文字檔案", "*.txt"), ("所有檔案", "*.*")]
        ))
        
        if not txt_files:
            print("未選擇 txt 檔案，程式結束。")
            messagebox.showinfo("提示", "未選擇 txt 檔案")
            return
        
        # 過濾掉包含 Timestamp 的檔案
        filtered_files = []
        for txt_file in txt_files:
            filename = os.path.basename(txt_file)
            if 'Timestamp' in filename or 'timestamp' in filename:
                skipped_timestamp_files.append(filename)
            else:
                filtered_files.append(txt_file)
        
        txt_files = filtered_files
        
        print(f"已選擇 {len(txt_files)} 個 txt 檔案")
        if skipped_timestamp_files:
            print(f"跳過 {len(skipped_timestamp_files)} 個 Timestamp 檔案")
    else:
        # 選擇資料夾並自動掃描
        print("\n請選擇包含 txt 檔案的資料夾...")
        folder_path = askdirectory(title="選擇包含 txt 檔案的資料夾")
        
        if not folder_path:
            print("未選擇資料夾，程式結束。")
            messagebox.showinfo("提示", "未選擇資料夾")
            return
        
        # 掃描資料夾中的 txt 檔案
        txt_files, skipped_timestamp_files = scan_folder_for_txt(folder_path)
        
        if not txt_files and not skipped_timestamp_files:
            print("資料夾中沒有找到 txt 檔案。")
            messagebox.showinfo("提示", "資料夾中沒有找到 txt 檔案")
            return
        
        print(f"\n在資料夾中找到:")
        print(f"  有效的 txt 檔案: {len(txt_files)} 個")
        if skipped_timestamp_files:
            print(f"  跳過的 Timestamp 檔案: {len(skipped_timestamp_files)} 個")
        
        if txt_files:
            print(f"\n待處理的檔案:")
            for txt_file in txt_files:
                print(f"  → {os.path.basename(txt_file)}")
        
        if skipped_timestamp_files:
            print(f"\n自動跳過的 Timestamp 檔案:")
            for filename in skipped_timestamp_files:
                print(f"  ✗ {filename}")
    
    if not txt_files:
        print("\n沒有可處理的 txt 檔案（所有檔案都包含 Timestamp）。")
        messagebox.showinfo("提示", "沒有可處理的 txt 檔案")
        return
    
    # 確認操作
    confirm_message = (
        f"即將處理：\n"
        f"• Excel 檔案: {os.path.basename(excel_path)}\n"
        f"• txt 檔案數量: {len(txt_files)}\n"
    )
    
    if skipped_timestamp_files:
        confirm_message += f"• 跳過 Timestamp 檔案: {len(skipped_timestamp_files)}\n"
    
    confirm_message += (
        f"\n填入規則：\n"
        f"• 分類（B欄）：檔名的第一個單字\n"
        f"• 標題（C欄）：完整檔名\n"
        f"• 內文（D欄）：txt 內容（自動換行）\n"
        f"• 大類（A欄）：保持空白\n"
        f"• 自動跳過：包含 'Timestamp' 的檔案\n\n"
        f"Excel 檔案將被直接覆蓋更新。\n"
        f"是否繼續？"
    )
    
    confirm = messagebox.askyesno("確認操作", confirm_message)
    
    if not confirm:
        print("使用者取消操作")
        return
    
    # 步驟 3: 處理檔案
    try:
        print("\n" + "=" * 60)
        print("開始處理...")
        print("=" * 60)
        
        success_count, skip_exist_count, skip_timestamp_count = fill_excel_with_txt(excel_path, txt_files)
        
        # 顯示結果
        print("\n" + "=" * 60)
        print("處理完成！")
        print(f"成功填入: {success_count} 個")
        print(f"跳過（已存在）: {skip_exist_count} 個")
        print(f"跳過（Timestamp）: {skip_timestamp_count + len(skipped_timestamp_files)} 個")
        print("=" * 60)
        
        result_message = (
            f"處理完成！\n\n"
            f"成功填入: {success_count} 個\n"
            f"跳過（已存在）: {skip_exist_count} 個\n"
        )
        
        if skip_timestamp_count + len(skipped_timestamp_files) > 0:
            result_message += f"跳過（Timestamp）: {skip_timestamp_count + len(skipped_timestamp_files)} 個\n"
        
        result_message += f"\nExcel 檔案已更新：\n{excel_path}"
        
        messagebox.showinfo("完成", result_message)
        
    except Exception as e:
        error_msg = f"處理過程發生錯誤：\n{str(e)}"
        print(f"\n錯誤: {error_msg}")
        messagebox.showerror("錯誤", error_msg)

if __name__ == "__main__":
    main()
