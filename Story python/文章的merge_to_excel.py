"""
merge_to_excel.py
-----------------
將從 Web 匯出的自訂文章 JSON 合併進 story.xlsx。

用法：
    python merge_to_excel.py custom-articles-2025-01-01.json story.xlsx

參數：
    json_file   : 從 Web 匯出的 JSON 檔案路徑
    excel_file  : 目標 Excel 檔案路徑（會直接修改，建議先備份）

規則：
    - 以 slug 欄位判斷重複：slug 相同的文章會跳過（不覆蓋）
    - 如果 Excel 不存在，會建立新檔
    - 新增的資料會 append 到現有資料的最後一列
"""

import sys
import json
import os
from datetime import datetime


def load_json(path):
    with open(path, encoding="utf-8") as f:
        data = json.load(f)
    if not isinstance(data, list):
        raise ValueError("JSON 必須是陣列 (list)")
    return data


def load_excel(path):
    try:
        import openpyxl
    except ImportError:
        print("❌ 請先安裝 openpyxl：pip install openpyxl")
        sys.exit(1)

    if os.path.exists(path):
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        print(f"✅ 已開啟現有 Excel：{path}（共 {ws.max_row - 1} 筆資料）")
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "工作表1"
        # 建立表頭（對應你的 Excel 欄位）
        ws.append(["大類", "分類", "標題", "內文", "slug"])
        print(f"📄 Excel 不存在，已建立新檔：{path}")

    return wb, ws


def get_existing_slugs(ws):
    """讀取第一列作為表頭，找到 slug 欄，收集所有現有 slug"""
    headers = [cell.value for cell in ws[1]]

    # 找 slug 欄的索引（1-based for openpyxl）
    slug_col = None
    for i, h in enumerate(headers):
        if h and str(h).strip().lower() == "slug":
            slug_col = i + 1  # openpyxl 是 1-based
            break

    if slug_col is None:
        # 如果沒有 slug 欄，就用第 5 欄（你的 Excel 第 5 欄是 slug）
        slug_col = 5
        print("⚠️  找不到 'slug' 表頭，預設使用第 5 欄作為 slug 欄")

    existing_slugs = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        val = row[slug_col - 1]  # 轉成 0-based
        if val:
            existing_slugs.add(str(val).strip().lower())

    return existing_slugs, slug_col, headers


def merge(json_path, excel_path):
    articles = load_json(json_path)
    wb, ws = load_excel(excel_path)
    existing_slugs, slug_col, headers = get_existing_slugs(ws)

    # 找各欄位的索引，優先對應表頭名稱，找不到就用預設順序
    def col_index(names, fallback):
        for name in names:
            for i, h in enumerate(headers):
                if h and str(h).strip() == name:
                    return i
        return fallback

    idx_major    = col_index(["大類"],  0)
    idx_category = col_index(["分類"],  1)
    idx_title    = col_index(["標題"],  2)
    idx_content  = col_index(["內文"],  3)
    idx_slug     = col_index(["slug"],  4)

    added = 0
    skipped = 0

    for art in articles:
        # 相容兩種格式（Web 匯出格式 / 內部格式）
        major    = art.get("大類")    or art.get("major")    or ""
        category = art.get("分類")    or art.get("category") or ""
        title    = art.get("標題")    or art.get("title")    or ""
        content  = art.get("內文")    or art.get("content")  or ""
        slug     = art.get("slug")    or art.get("slug")     or ""

        if not title or not content:
            print(f"  ⏭  跳過（缺少標題或內文）：{art}")
            skipped += 1
            continue

        if slug.lower() in existing_slugs:
            print(f"  ⏭  已存在，跳過：{slug}")
            skipped += 1
            continue

        # 建立新列，欄數對應現有表頭數量
        num_cols = max(len(headers), 5)
        new_row = [""] * num_cols
        new_row[idx_major]    = major
        new_row[idx_category] = category
        new_row[idx_title]    = title
        new_row[idx_content]  = content
        new_row[idx_slug]     = slug

        ws.append(new_row)
        existing_slugs.add(slug.lower())
        added += 1
        print(f"  ✅ 新增：{title} ({slug})")

    wb.save(excel_path)

    print(f"\n{'='*50}")
    print(f"完成！新增 {added} 筆，跳過 {skipped} 筆")
    print(f"已儲存至：{excel_path}")
    print(f"{'='*50}")


def main():
    if len(sys.argv) < 3:
        print(__doc__)
        print("\n錯誤：請提供兩個參數")
        print("用法：python merge_to_excel.py <json檔> <excel檔>")
        sys.exit(1)

    json_path  = sys.argv[1]
    excel_path = sys.argv[2]

    if not os.path.exists(json_path):
        print(f"❌ 找不到 JSON 檔：{json_path}")
        sys.exit(1)

    # 自動備份 Excel（如果存在）
    if os.path.exists(excel_path):
        backup = excel_path.replace(".xlsx", f"_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
        import shutil
        shutil.copy2(excel_path, backup)
        print(f"💾 已自動備份 Excel → {backup}")

    merge(json_path, excel_path)


if __name__ == "__main__":
    main()
